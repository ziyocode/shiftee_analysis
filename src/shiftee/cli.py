#!/usr/bin/env python3
"""
Shiftee 근무 데이터 분석 및 초과근로 적정성 판정 CLI 도구

Usage:
    shiftee-analyze [OPTIONS]
"""

import sys
import asyncio
from pathlib import Path
from datetime import datetime, timedelta
import re
import argparse
import pandas as pd
import openpyxl

# shiftee 패키지 내부 모듈
from .settings import ShifteeSettings
from .login import launch_browser, login
from .attendance import download_report_current_month, download_payroll_current_month

# kakao_send 모듈 (src/kakao_send)
try:
    from kakao_send import Kakao, format_risk_message
    from kakao_send.message_formatter import format_summary_message
    KAKAO_AVAILABLE = True
except ImportError:
    KAKAO_AVAILABLE = False


async def download_shiftee_data(start_date: datetime, end_date: datetime, output_dir: Path):
    """Shiftee에서 데이터 다운로드."""
    print("\n" + "=" * 80)
    print("📥 Shiftee 데이터 다운로드")
    print("=" * 80)

    start_str = start_date.strftime("%Y.%m.%d")
    end_str = end_date.strftime("%Y.%m.%d")
    print(f"기간: {start_str} ~ {end_str}\n")

    settings = ShifteeSettings()

    async with launch_browser(settings) as (browser, context, page):
        print("🔐 로그인 중...")
        await login(page, settings)
        print("   ✅ 로그인 완료\n")

        print("📊 1. 리포트 다운로드 중...")
        data1_path = await download_report_current_month(
            page, settings, output_dir, start_str, end_str
        )
        print(f"   ✅ 저장: {data1_path}\n")

        print("💰 2. 급여정산 다운로드 중...")
        data2_path = await download_payroll_current_month(
            page, settings, output_dir, start_str, end_str
        )
        print(f"   ✅ 저장: {data2_path}\n")

        return data1_path, data2_path


def find_date_range_sheet(wb: openpyxl.Workbook) -> str:
    """날짜 범위 패턴(YYYYMMDD-YYYYMMDD)의 시트 찾기."""
    pattern = re.compile(r"\d{8}-\d{8}")
    for sheet_name in wb.sheetnames:
        if pattern.match(sheet_name):
            return sheet_name

    raise ValueError(
        f"날짜 범위 시트(YYYYMMDD-YYYYMMDD)를 찾을 수 없습니다. "
        f"사용 가능한 시트: {wb.sheetnames}"
    )


def load_shiftee_data1(file_path: Path) -> pd.DataFrame:
    """shiftee_data1.xlsx (REALTIME-REPORT) 로드."""
    print(f"📂 {file_path.name} 로드 중...")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet_name = find_date_range_sheet(wb)
    print(f"   ✅ 시트 발견: {sheet_name}")

    sheet = wb[sheet_name]
    headers = []
    for col_idx in range(1, sheet.max_column + 1):
        header = sheet.cell(1, col_idx).value
        headers.append(header if header else f"Column_{col_idx}")

    data = []
    for row_idx in range(2, sheet.max_row + 1):
        row_data = []
        for col_idx in range(1, len(headers) + 1):
            row_data.append(sheet.cell(row_idx, col_idx).value)
        data.append(row_data)

    wb.close()
    df = pd.DataFrame(data, columns=headers)
    print(f"   📊 데이터: {len(df)}행 × {len(df.columns)}열\n")
    return df


def load_shiftee_data2(file_path: Path) -> pd.DataFrame:
    """shiftee_data2.xlsx (PAYROLL) 로드."""
    print(f"📂 {file_path.name} 로드 중...")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active

    headers = []
    for col_idx in range(1, sheet.max_column + 1):
        header = sheet.cell(3, col_idx).value
        headers.append(header if header else f"Column_{col_idx}")

    data = []
    for row_idx in range(4, sheet.max_row + 1):
        row_data = []
        for col_idx in range(1, len(headers) + 1):
            row_data.append(sheet.cell(row_idx, col_idx).value)
        data.append(row_data)

    wb.close()
    df = pd.DataFrame(data, columns=headers)
    print(f"   📊 데이터: {len(df)}행 × {len(df.columns)}열\n")
    return df


def calculate_basic_columns(df1: pd.DataFrame) -> pd.DataFrame:
    """기본 열 계산."""
    print("🔢 기본 열 계산 중...")
    df = pd.DataFrame()
    df["B_직원"] = df1["직원"]
    df["C_본조직"] = df1["본조직"]
    df["D_소정근로시간"] = pd.to_numeric(df1["소정근로시간"], errors='coerce')
    df["E_승인된근로시간"] = pd.to_numeric(df1["승인된 근로시간"], errors='coerce')
    df["F_실제근로시간"] = pd.to_numeric(df1["실제 근로시간"], errors='coerce')
    df["K_유급휴가시간"] = pd.to_numeric(df1["유급휴가시간"], errors='coerce')
    df["표준근로시간"] = pd.to_numeric(df1["표준 근로시간"], errors='coerce')
    df["표준최대잔여유급시간"] = pd.to_numeric(df1.get("표준 최대 잔여유급시간", 0), errors='coerce').fillna(0)
    df["결근"] = pd.to_numeric(df1["결근"], errors='coerce').fillna(0)
    df["퇴근누락"] = pd.to_numeric(df1["퇴근 누락"], errors='coerce').fillna(0)
    print(f"   ✅ 기본 열 계산 완료\n")
    return df


def calculate_g_column(df: pd.DataFrame) -> pd.DataFrame:
    """G열 계산: 실제 근로시간(결근,퇴근누락포함)."""
    print("🔢 G열 계산 중 (실제 근로시간, 결근/퇴근누락포함)...")
    df["G_실제근로시간_결근포함"] = (
        df["F_실제근로시간"] + df["결근"] * 8 + df["퇴근누락"] * 8
    ).round(2)
    print(f"   ✅ G열 계산 완료\n")
    return df


def calculate_h_column(df: pd.DataFrame, df1: pd.DataFrame, df2: pd.DataFrame) -> pd.DataFrame:
    """H열 계산: 실제 근로시간(결근,퇴근누락포함) 실제퇴근시간-출근등록시간."""
    print("🔢 H열 계산 중 (SUMPRODUCT 로직, 실제퇴근-출근)...")
    from datetime import datetime, time, timedelta

    df2_filtered = df2[["이름", "근무일정\n시작시간", "퇴근시간", "(실제)\n총 휴게시간"]].copy()
    df2_filtered.columns = ["직원명", "시작시간", "퇴근시간", "휴게시간"]
    df2_filtered = df2_filtered[df2_filtered["퇴근시간"].notna()]

    def calculate_work_hours(row):
        try:
            start, end, rest = row["시작시간"], row["퇴근시간"], row["휴게시간"]

            def to_excel_serial(dt_value):
                if isinstance(dt_value, datetime):
                    excel_epoch = datetime(1899, 12, 30)
                    return (dt_value - excel_epoch).total_seconds() / 86400
                elif isinstance(dt_value, (int, float)):
                    return dt_value
                return 0

            start_serial = to_excel_serial(start)
            end_serial = to_excel_serial(end)
            start_day = int(start_serial * 1440) / 1440
            end_day = int(end_serial * 1440) / 1440

            rest_hours = 0
            if rest is not None:
                if isinstance(rest, time):
                    rest_hours = rest.hour + rest.minute / 60 + rest.second / 3600
                elif isinstance(rest, timedelta):
                    rest_hours = rest.total_seconds() / 3600
                elif isinstance(rest, (int, float)):
                    rest_hours = rest * 24
            
            return max(0, (end_day - start_day) * 24 - rest_hours)
        except Exception:
            return 0

    df2_filtered["근무시간"] = df2_filtered.apply(calculate_work_hours, axis=1)
    employee_hours = df2_filtered.groupby("직원명")["근무시간"].sum().to_dict()

    def get_h_value(row):
        return employee_hours.get(row["B_직원"], 0) + row["결근"] * 8 + row["퇴근누락"] * 8

    df["H_실제근로시간_퇴근출근기반"] = df.apply(get_h_value, axis=1).round(2)
    print(f"   ✅ H열 계산 완료 (SUMPRODUCT)\n")
    return df


def calculate_i_j_l_columns(df: pd.DataFrame) -> pd.DataFrame:
    """I, J, L열 계산."""
    print("🔢 I, J, L열 계산 중...")
    df["I_표준근로시간"] = (df["표준근로시간"] + df["표준최대잔여유급시간"]).round(2)
    df["J_표준근로시간_결근포함"] = (
        df["표준근로시간"] + df["결근"] * 8 + df["퇴근누락"] * 8 + df["표준최대잔여유급시간"]
    ).round(2)
    df["L_법정근로시간"] = (df["D_소정근로시간"] - df["K_유급휴가시간"]).round(2)
    print(f"   ✅ I, J, L열 계산 완료\n")
    return df


def calculate_overtime_columns(df: pd.DataFrame) -> pd.DataFrame:
    """초과근로 관련 열 계산."""
    print("🔢 초과근로 열 계산 중 (M, N, O, P)...")
    df["M_실제초과근로시간"] = df.apply(
        lambda row: max(0, row["F_실제근로시간"] - row["L_법정근로시간"]), axis=1
    ).round(2)
    df["N_실제초과근로_결근포함"] = df.apply(
        lambda row: max(0, row["G_실제근로시간_결근포함"] - row["L_법정근로시간"]), axis=1
    ).round(2)

    def calculate_o(row):
        h_minus_l = row["H_실제근로시간_퇴근출근기반"] - row["L_법정근로시간"]
        n, h = row["N_실제초과근로_결근포함"], row["H_실제근로시간_퇴근출근기반"]
        if h_minus_l > n or h > 300:
            return n
        return 0 if h_minus_l < 0 else h_minus_l

    df["O_실제초과근로_조기출근제외"] = df.apply(calculate_o, axis=1).round(2)
    df["P_조기출근합산"] = (df["N_실제초과근로_결근포함"] - df["O_실제초과근로_조기출근제외"]).round(2)
    print(f"   ✅ 초과근로 열 계산 완료\n")
    return df


def calculate_legal_limits_and_risk(df: pd.DataFrame, start_date: datetime, end_date: datetime) -> pd.DataFrame:
    """법규 기준 및 적정성 판정."""
    print(f"🔢 법규 기준 및 적정성 판정 중 ({start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')})...")
    last_day = end_date.day
    q_value = round(last_day / 7 * 10, 2)
    r_value = round(last_day / 7 * 12, 2)
    s_value = round(12 * 4.3, 2)
    
    print(f"   📅 기준일: {last_day}일, 법정초과: {q_value}h, 법규위반: {r_value}h")

    df["Q_법정초과근로시간"] = q_value
    df["R_법규위반_전일까지"] = r_value
    df["S_월법규위반시간"] = s_value

    df["T_월말가능초과근로"] = df.apply(
        lambda row: (
            "가능시간없음"
            if (row["S_월법규위반시간"] - row["O_실제초과근로_조기출근제외"]) < 0
            else round(row["S_월법규위반시간"] - row["O_실제초과근로_조기출근제외"], 2)
        ), axis=1
    )
    df["U_적정성"] = df.apply(
        lambda row: "위험" if row["O_실제초과근로_조기출근제외"] > row["Q_법정초과근로시간"] else "정상", axis=1
    )
    df["V_법규기준초과자"] = df.apply(
        lambda row: "법기준초과" if row["O_실제초과근로_조기출근제외"] != 0 and row["O_실제초과근로_조기출근제외"] >= row["R_법규위반_전일까지"] else "",
        axis=1
    )
    print(f"   ✅ 법규 기준 및 적정성 판정 완료\n")
    return df


def print_summary(df: pd.DataFrame):
    """계산 결과 요약 출력."""
    print("\n" + "=" * 80 + "\n📊 적정성 분석 결과\n" + "=" * 80)
    total = len(df)
    risk_count = len(df[df["U_적정성"] == "위험"])
    legal_exceed_count = len(df[df["V_법규기준초과자"] == "법기준초과"])
    print(f"\n총 직원: {total}명")
    print(f"  - ✅ 정상: {total - risk_count}명")
    print(f"  - ⚠️  위험: {risk_count}명")
    print(f"  - 🚨 법규기준초과: {legal_exceed_count}명")


def print_risk_employees(df: pd.DataFrame):
    """위험 직원 목록 출력."""
    risk_df = df[df["U_적정성"] == "위험"].copy()
    if risk_df.empty:
        print("\n✅ 위험 직원이 없습니다!\n")
        return

    print("\n" + "=" * 80 + "\n⚠️  위험 직원 목록\n" + "=" * 80)
    display_df = risk_df[["B_직원", "C_본조직", "O_실제초과근로_조기출근제외", "Q_법정초과근로시간", "U_적정성", "V_법규기준초과자"]].copy()
    display_df.columns = ["직원", "본조직", "실제초과근로(h)", "법정기준(h)", "적정성", "법규초과"]
    print(display_df.sort_values("실제초과근로(h)", ascending=False).to_string(index=False))
    print()


def save_to_excel(df: pd.DataFrame, output_path: Path, start_date: datetime, end_date: datetime):
    """결과를 Excel 파일로 저장하고 스타일링 적용."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    print(f"💾 Excel 저장 중... {output_path}")

    column_mapping = {
        "B_직원": "직원", "C_본조직": "본조직", "D_소정근로시간": "소정\n근로시간",
        "E_승인된근로시간": "승인된\n근로시간", "F_실제근로시간": "실제\n근로시간",
        "G_실제근로시간_결근포함": "실제\n근로시간\n(결근,퇴근누락\n포함)",
        "H_실제근로시간_퇴근출근기반": "실제\n근로시간\n(퇴근출근기반)", "I_표준근로시간": "표준\n근로시간",
        "J_표준근로시간_결근포함": "표준\n근로시간\n(결근포함)", "K_유급휴가시간": "유급휴가\n시간",
        "L_법정근로시간": "법정\n근로시간", "M_실제초과근로시간": "실제 초과\n근로시간",
        "N_실제초과근로_결근포함": "실제 초과\n근로시간\n(결근포함)",
        "O_실제초과근로_조기출근제외": "실제 초과\n근로시간\n(조기출근제외)",
        "P_조기출근합산": "조기출근\n합산", "Q_법정초과근로시간": "법정\n초과 근로시간",
        "R_법규위반_전일까지": "법규 위반\n(전일까지)", "S_월법규위반시간": "월법규\n위반시간",
        "T_월말가능초과근로": "월말까지\n가능한\n초과근로시간", "U_적정성": "적정성",
        "V_법규기준초과자": "법규 기준초과자",
    }
    
    display_df = df[list(column_mapping.keys())].copy()
    display_df = display_df.rename(columns=column_mapping)
    display_df.insert(0, "", "")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        display_df.to_excel(writer, sheet_name="계산", index=False)
        ws = writer.sheets["계산"]
        
        # 스타일링 로직 (간소화)
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        risk_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        legal_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
        
        for col_idx in range(1, len(display_df.columns) + 1):
            ws.cell(row=1, column=col_idx).fill = header_fill

        adequacy_idx = display_df.columns.get_loc("적정성") + 1
        legal_idx = display_df.columns.get_loc("법규 기준초과자") + 1

        for row_idx in range(2, len(display_df) + 2):
            if ws.cell(row=row_idx, column=adequacy_idx).value == "위험":
                for c in range(1, len(display_df.columns) + 1):
                    ws.cell(row=row_idx, column=c).fill = risk_fill
            elif ws.cell(row=row_idx, column=legal_idx).value == "법기준초과":
                for c in range(1, len(display_df.columns) + 1):
                    ws.cell(row=row_idx, column=c).fill = legal_fill

        ws.freeze_panes = "B2"
        # 열 너비 자동 조정 생략 (기본 너비 사용)

    print(f"📊 Excel 리포트 생성 완료: {output_path} ({output_path.stat().st_size / 1024:.1f} KB)\n")


def main():
    """메인 함수."""
    parser = argparse.ArgumentParser(description="Shiftee 근무 데이터 분석 및 적정성 판정")
    parser.add_argument("--data1", type=Path, default=Path("data/shiftee_data1.xlsx"))
    parser.add_argument("--data2", type=Path, default=Path("data/shiftee_data2.xlsx"))
    parser.add_argument("--output", "-o", type=Path)
    parser.add_argument("--start", type=str)
    parser.add_argument("--end", type=str)
    parser.add_argument("--download", action="store_true")
    parser.add_argument("--send-kakao", action="store_true")
    parser.add_argument("--kakao-summary", action="store_true")
    
    args = parser.parse_args()

    # 날짜 기본값 설정
    now = datetime.now()
    start_date = datetime.strptime(args.start, "%Y-%m-%d") if args.start else datetime(now.year, now.month, 1)
    end_date = datetime.strptime(args.end, "%Y-%m-%d") if args.end else (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    
    if not args.output:
        args.output = Path("output") / f"report_{datetime.now().strftime('%Y%m%d')}.xlsx"

    try:
        if args.download:
            args.data1.parent.mkdir(parents=True, exist_ok=True)
            args.data1, args.data2 = asyncio.run(download_shiftee_data(start_date, end_date, args.data1.parent))

        print("=" * 80 + "\n🚀 Shiftee 적정성 위험 판정\n" + "=" * 80 + "\n")

        print("📂 1단계: 데이터 로드")
        df1 = load_shiftee_data1(args.data1)
        df2 = load_shiftee_data2(args.data2)

        # 교대제 제외
        if "본직무" in df1.columns:
            shift_workers = df1[df1["본직무"] == "교대제"]["직원"].tolist() if "직원" in df1.columns else []
            df1 = df1[df1["본직무"] != "교대제"].copy()
            if shift_workers:
                print(f"⚠️  교대제 직원 {len(shift_workers)}명 제외: {', '.join(shift_workers[:5])}...")
                if "이름" in df2.columns:
                    df2 = df2[~df2["이름"].isin(shift_workers)].copy()

        print("🔢 2단계: 계산 수행")
        df = calculate_basic_columns(df1)
        df = calculate_g_column(df)
        df = calculate_h_column(df, df1, df2)
        df = calculate_i_j_l_columns(df)
        df = calculate_overtime_columns(df)
        df = calculate_legal_limits_and_risk(df, start_date, end_date)

        print_summary(df)
        print_risk_employees(df)

        print("💾 결과 저장 중...")
        args.output.parent.mkdir(parents=True, exist_ok=True)
        if args.output.suffix.lower() == ".xlsx":
            save_to_excel(df, args.output, start_date, end_date)
        else:
            df.to_csv(args.output, index=False, encoding="utf-8-sig")
            print(f"💾 CSV 저장 완료: {args.output}\n")

        if args.send_kakao or args.kakao_summary:
            if not KAKAO_AVAILABLE:
                print("\n⚠️  kakao_send 모듈을 불러올 수 없습니다.")
                return 1
                
            print("📱 카카오톡 메시지 전송")
            kakao = Kakao()
            msg = format_summary_message(df, start_date, end_date) if args.kakao_summary else format_risk_message(df, start_date, end_date, show_all=False)
            if kakao.send_message(msg):
                print("   ✅ 전송 완료")
            else:
                print("   ⚠️  전송 실패")

    except ValueError as e:
         print(f"❌ 값 오류: {e}")
         return 1
    except FileNotFoundError as e:
         print(f"❌ 파일 없음: {e}")
         return 1
    except Exception as e:
         print(f"❌ 오류 발생: {e}")
         import traceback
         traceback.print_exc()
         return 1
    
    return 0


if __name__ == "__main__":
    sys.exit(main())
