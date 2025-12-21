"""Excel 파일 처리 모듈.

openpyxl을 사용하여 수식을 보존하면서 Excel 파일을 읽고 쓰는 기능 제공.
"""

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelProcessor:
    """Excel 파일 처리 클래스."""

    def __init__(self, filepath: Path | str):
        """Excel 파일 로드.

        Args:
            filepath: Excel 파일 경로
        """
        self.filepath = Path(filepath)
        if not self.filepath.exists():
            raise FileNotFoundError(f"Excel 파일을 찾을 수 없습니다: {self.filepath}")

        # openpyxl로 수식 보존하며 로드
        self.workbook: Workbook = load_workbook(
            filename=str(self.filepath),
            data_only=False,  # 수식 보존
            keep_vba=False,  # VBA 제거 - macOS Excel 호환성 향상
            rich_text=True,  # 서식있는 텍스트 보존
        )

    def get_sheet_names(self) -> list[str]:
        """모든 시트 이름 반환.

        Returns:
            시트 이름 리스트
        """
        return self.workbook.sheetnames

    def get_sheet(self, sheet_name: str) -> Worksheet:
        """특정 시트 반환.

        Args:
            sheet_name: 시트 이름

        Returns:
            Worksheet 객체

        Raises:
            ValueError: 시트가 존재하지 않을 때
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(
                f"시트 '{sheet_name}'를 찾을 수 없습니다. "
                f"사용 가능한 시트: {self.workbook.sheetnames}"
            )
        return self.workbook[sheet_name]

    def read_as_dataframe(
        self, sheet_name: str, header_row: int = 0, skip_rows: int = 0
    ) -> pd.DataFrame:
        """시트를 pandas DataFrame으로 읽기.

        Args:
            sheet_name: 시트 이름
            header_row: 헤더 행 번호 (0-based)
            skip_rows: 건너뛸 행 수

        Returns:
            DataFrame
        """
        # pandas로 읽기 (데이터만, 수식은 계산된 값)
        df = pd.read_excel(
            self.filepath,
            sheet_name=sheet_name,
            header=header_row,
            skiprows=skip_rows if skip_rows > 0 else None,
        )
        return df

    def write_dataframe_to_sheet(
        self,
        df: pd.DataFrame,
        sheet_name: str,
        start_row: int = 1,
        start_col: int = 1,
        include_header: bool = True,
        clear_existing: bool = True,
    ) -> None:
        """DataFrame을 시트에 쓰기 (기존 수식 보존).

        Args:
            df: 쓸 DataFrame
            sheet_name: 대상 시트 이름
            start_row: 시작 행 (1-based)
            start_col: 시작 열 (1-based)
            include_header: 헤더 포함 여부
            clear_existing: 기존 데이터 삭제 여부
        """
        sheet = self.get_sheet(sheet_name)

        # 기존 데이터 삭제 (수식은 유지하고 값만 삭제)
        if clear_existing:
            self._clear_data_range(
                sheet,
                start_row,
                start_col,
                start_row + len(df) + (1 if include_header else 0),
                start_col + len(df.columns),
            )

        # 헤더 쓰기
        if include_header:
            for col_idx, col_name in enumerate(df.columns):
                cell = sheet.cell(row=start_row, column=start_col + col_idx)
                cell.value = col_name
            start_row += 1

        # 데이터 쓰기
        for row_idx, row_data in enumerate(df.itertuples(index=False)):
            for col_idx, value in enumerate(row_data):
                cell = sheet.cell(row=start_row + row_idx, column=start_col + col_idx)
                # NaN이나 None은 빈 셀로
                if pd.isna(value):
                    cell.value = None
                else:
                    cell.value = value

    def _clear_data_range(
        self,
        sheet: Worksheet,
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int,
    ) -> None:
        """지정된 범위의 데이터를 삭제 (수식이 아닌 값만).

        Args:
            sheet: 대상 시트
            start_row: 시작 행 (1-based)
            start_col: 시작 열 (1-based)
            end_row: 끝 행 (1-based)
            end_col: 끝 열 (1-based)
        """
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = sheet.cell(row=row, column=col)
                # 수식이 없는 셀만 삭제
                if not (cell.value and str(cell.value).startswith("=")):
                    cell.value = None

    def clear_sheet_data(
        self, sheet_name: str, start_row: int = 2, preserve_formulas: bool = True
    ) -> None:
        """시트의 데이터를 삭제 (헤더와 수식은 보존).

        Args:
            sheet_name: 시트 이름
            start_row: 데이터 시작 행 (1-based, 기본값 2는 헤더 다음)
            preserve_formulas: 수식 보존 여부
        """
        sheet = self.get_sheet(sheet_name)

        for row in sheet.iter_rows(min_row=start_row):
            for cell in row:
                if preserve_formulas:
                    # 수식이 없는 셀만 삭제
                    if not (cell.value and str(cell.value).startswith("=")):
                        cell.value = None
                else:
                    # 모든 셀 삭제
                    cell.value = None

    def enable_formula_recalculation(self) -> None:
        """Excel 파일을 열 때 수식 자동 재계산 활성화."""
        # 전체 workbook에 대해 재계산 모드 설정
        self.workbook.calculation.calcMode = "auto"
        self.workbook.calculation.fullCalcOnLoad = True

    def _create_fresh_workbook(self) -> Workbook:
        """Excel 호환성을 위해 완전히 새로운 Workbook 생성.

        기존 workbook의 데이터를 새 workbook으로 복사하여
        템플릿의 메타데이터나 호환성 문제를 제거합니다.

        Returns:
            새로 생성된 Workbook
        """
        from openpyxl import Workbook

        # 완전히 새로운 Workbook 생성
        new_wb = Workbook()
        new_wb.remove(new_wb.active)  # 기본 시트 제거

        # 각 시트 복사
        for sheet_name in self.workbook.sheetnames:
            source_sheet = self.workbook[sheet_name]
            target_sheet = new_wb.create_sheet(sheet_name)

            # 모든 셀 값 복사 (수식 포함)
            for row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row):
                for cell in row:
                    if cell.value is not None:
                        target_cell = target_sheet[cell.coordinate]
                        target_cell.value = cell.value

            # 기본 시트 속성 복사
            target_sheet.sheet_properties.tabColor = source_sheet.sheet_properties.tabColor

        return new_wb

    def save(self, output_path: Path | str | None = None, recalculate: bool = True,
             use_fresh_workbook: bool = True) -> Path:
        """Workbook을 파일로 저장.

        Args:
            output_path: 저장 경로 (None이면 원본 파일에 덮어쓰기)
            recalculate: 수식 자동 재계산 활성화 여부
            use_fresh_workbook: Excel 호환성을 위해 새 Workbook으로 복사하여 저장 (권장)

        Returns:
            저장된 파일 경로
        """
        if output_path is None:
            output_path = self.filepath
        else:
            output_path = Path(output_path)

        # 디렉토리 생성
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Excel 호환성을 위해 새 Workbook 생성 (기본값)
        if use_fresh_workbook:
            workbook_to_save = self._create_fresh_workbook()
        else:
            workbook_to_save = self.workbook

        # 수식 재계산 활성화
        if recalculate:
            workbook_to_save.calculation.calcMode = "auto"
            workbook_to_save.calculation.fullCalcOnLoad = True

        # calcChain 제거 - Excel 호환성 문제 해결
        if hasattr(workbook_to_save, '_calcChain'):
            workbook_to_save._calcChain = None

        # 외부 링크 제거 - 깨진 외부 참조로 인한 Excel 오류 해결
        if hasattr(workbook_to_save, '_external_links'):
            workbook_to_save._external_links = []

        # 저장
        workbook_to_save.save(str(output_path))

        # 새로 생성한 workbook 정리
        if use_fresh_workbook:
            workbook_to_save.close()

        return output_path

    def copy_to(self, output_path: Path | str) -> "ExcelProcessor":
        """현재 workbook을 새 파일로 복사하여 새 ExcelProcessor 반환.

        Args:
            output_path: 복사할 경로

        Returns:
            새로운 ExcelProcessor 인스턴스
        """
        output_path = Path(output_path)
        self.save(output_path)
        return ExcelProcessor(output_path)

    def close(self) -> None:
        """Workbook 닫기."""
        self.workbook.close()

    def __enter__(self) -> "ExcelProcessor":
        """Context manager 진입."""
        return self

    def __exit__(self, exc_type: Any, exc_val: Any, exc_tb: Any) -> None:
        """Context manager 종료."""
        self.close()


def load_excel(filepath: Path | str) -> ExcelProcessor:
    """Excel 파일 로드 (헬퍼 함수).

    Args:
        filepath: Excel 파일 경로

    Returns:
        ExcelProcessor 인스턴스
    """
    return ExcelProcessor(filepath)


def read_excel_to_dataframe(
    filepath: Path | str, sheet_name: str, header_row: int = 0
) -> pd.DataFrame:
    """Excel 시트를 DataFrame으로 읽기 (헬퍼 함수).

    Args:
        filepath: Excel 파일 경로
        sheet_name: 시트 이름
        header_row: 헤더 행 번호 (0-based)

    Returns:
        DataFrame
    """
    with ExcelProcessor(filepath) as processor:
        return processor.read_as_dataframe(sheet_name, header_row=header_row)
