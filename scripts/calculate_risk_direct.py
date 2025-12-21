#!/usr/bin/env python3
"""
Template 없이 직접 계산하는 적정성 위험 판정 스크립트

shiftee_data1.xlsx와 shiftee_data2.xlsx만으로 직접 계산하여
적정성이 "위험"인 직원을 찾습니다.

Usage:
    python scripts/calculate_risk_direct.py
    python scripts/calculate_risk_direct.py --data1 data/shiftee_data1.xlsx --data2 data/shiftee_data2.xlsx
    python scripts/calculate_risk_direct.py --output risk_report.csv
    python scripts/calculate_risk_direct.py --download --start 2025-11-01 --end 2025-11-30
    python scripts/calculate_risk_direct.py --send-kakao  # 위험 직원 목록을 카카오톡으로 전송
    python scripts/calculate_risk_direct.py --kakao-summary  # 요약만 카카오톡으로 전송
"""

import sys
import asyncio
from pathlib import Path
from datetime import datetime, timedelta
import calendar
import re

# 프로젝트 루트를 sys.path에 추가 (src 모듈을 찾을 수 있도록)
project_root = Path(__file__).parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

import pandas as pd
import openpyxl
import numpy as np


async def download_shiftee_data(start_date: datetime, end_date: datetime, output_dir: Path):
    """Shiftee에서 데이터 다운로드.

    Args:
        start_date: 시작 날짜
        end_date: 종료 날짜
        output_dir: 출력 디렉토리

    Returns:
        Tuple[Path, Path]: (data1_path, data2_path)
    """
    from src.shiftee.settings import ShifteeSettings
    from src.shiftee.login import launch_browser, login
    from src.shiftee.attendance import download_report_current_month, download_payroll_current_month

    print("\n" + "=" * 80)
    print("📥 Shiftee 데이터 다운로드")
    print("=" * 80)

    # 날짜 형식 변환 (YYYY-MM-DD → YYYY.MM.DD)
    start_str = start_date.strftime("%Y.%m.%d")
    end_str = end_date.strftime("%Y.%m.%d")
    print(f"기간: {start_str} ~ {end_str}\n")

    settings = ShifteeSettings()

    async with launch_browser(settings) as (browser, context, page):
        # 로그인
        print("🔐 로그인 중...")
        await login(page, settings)
        print("   ✅ 로그인 완료\n")

        # 1. 리포트 다운로드 (shiftee_data1.xlsx)
        print("📊 1. 리포트 다운로드 중...")
        data1_path = await download_report_current_month(
            page, settings, output_dir, start_str, end_str
        )
        print(f"   ✅ 저장: {data1_path}\n")

        # 2. 급여정산 다운로드 (shiftee_data2.xlsx)
        print("💰 2. 급여정산 다운로드 중...")
        data2_path = await download_payroll_current_month(
            page, settings, output_dir, start_str, end_str
        )
        print(f"   ✅ 저장: {data2_path}\n")

        return data1_path, data2_path


def find_date_range_sheet(wb: openpyxl.Workbook) -> str:
    """날짜 범위 패턴(YYYYMMDD-YYYYMMDD)의 시트 찾기.

    Args:
        wb: openpyxl Workbook 인스턴스

    Returns:
        시트 이름

    Raises:
        ValueError: 날짜 범위 시트를 찾지 못한 경우
    """
    pattern = re.compile(r"\d{8}-\d{8}")
    for sheet_name in wb.sheetnames:
        if pattern.match(sheet_name):
            return sheet_name

    raise ValueError(
        f"날짜 범위 시트(YYYYMMDD-YYYYMMDD)를 찾을 수 없습니다. "
        f"사용 가능한 시트: {wb.sheetnames}"
    )


def load_shiftee_data1(file_path: Path) -> pd.DataFrame:
    """shiftee_data1.xlsx (REALTIME-REPORT) 로드.

    Args:
        file_path: shiftee_data1.xlsx 파일 경로

    Returns:
        DataFrame with columns from the date range sheet
    """
    print(f"📂 {file_path.name} 로드 중...")

    wb = openpyxl.load_workbook(file_path, data_only=True)

    # 날짜 범위 시트 찾기
    sheet_name = find_date_range_sheet(wb)
    print(f"   ✅ 시트 발견: {sheet_name}")

    sheet = wb[sheet_name]

    # 헤더 추출 (1행)
    headers = []
    for col_idx in range(1, sheet.max_column + 1):
        header = sheet.cell(1, col_idx).value
        if header:
            headers.append(header)
        else:
            headers.append(f"Column_{col_idx}")

    # 데이터 추출 (2행부터)
    data = []
    for row_idx in range(2, sheet.max_row + 1):
        row_data = []
        for col_idx in range(1, len(headers) + 1):
            value = sheet.cell(row_idx, col_idx).value
            row_data.append(value)
        data.append(row_data)

    wb.close()

    df = pd.DataFrame(data, columns=headers)
    print(f"   📊 데이터: {len(df)}행 × {len(df.columns)}열\n")

    return df


def load_shiftee_data2(file_path: Path) -> pd.DataFrame:
    """shiftee_data2.xlsx (PAYROLL) 로드.

    Args:
        file_path: shiftee_data2.xlsx 파일 경로

    Returns:
        DataFrame with columns from the payroll sheet (header at row 3)
    """
    print(f"📂 {file_path.name} 로드 중...")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active

    # 헤더 추출 (3행)
    headers = []
    for col_idx in range(1, sheet.max_column + 1):
        header = sheet.cell(3, col_idx).value
        if header:
            headers.append(header)
        else:
            headers.append(f"Column_{col_idx}")

    # 데이터 추출 (4행부터)
    data = []
    for row_idx in range(4, sheet.max_row + 1):
        row_data = []
        for col_idx in range(1, len(headers) + 1):
            value = sheet.cell(row_idx, col_idx).value
            row_data.append(value)
        data.append(row_data)

    wb.close()

    df = pd.DataFrame(data, columns=headers)
    print(f"   📊 데이터: {len(df)}행 × {len(df.columns)}열\n")

    return df


def calculate_basic_columns(df1: pd.DataFrame) -> pd.DataFrame:
    """기본 열 계산 (D~K열: 원본 데이터 복사).

    Args:
        df1: shiftee_data1 DataFrame

    Returns:
        계산 결과 DataFrame
    """
    print("🔢 기본 열 계산 중...")

    df = pd.DataFrame()

    # B, C열: 직원, 본조직
    df["B_직원"] = df1["직원"]
    df["C_본조직"] = df1["본조직"]

    # D열: 소정근로시간 (원본 G열) - 숫자로 변환
    df["D_소정근로시간"] = pd.to_numeric(df1["소정근로시간"], errors='coerce')

    # E열: 승인된 근로시간 (원본 H열) - 숫자로 변환
    df["E_승인된근로시간"] = pd.to_numeric(df1["승인된 근로시간"], errors='coerce')

    # F열: 실제 근로시간 (원본 J열) - 숫자로 변환
    df["F_실제근로시간"] = pd.to_numeric(df1["실제 근로시간"], errors='coerce')

    # K열: 유급휴가시간 (원본 L열) - 숫자로 변환
    df["K_유급휴가시간"] = pd.to_numeric(df1["유급휴가시간"], errors='coerce')

    # 추가 데이터 (계산에 필요) - 모두 숫자로 변환
    df["표준근로시간"] = pd.to_numeric(df1["표준 근로시간"], errors='coerce')
    df["표준최대잔여유급시간"] = pd.to_numeric(df1.get("표준 최대 잔여유급시간", 0), errors='coerce').fillna(0)
    df["결근"] = pd.to_numeric(df1["결근"], errors='coerce').fillna(0)
    df["퇴근누락"] = pd.to_numeric(df1["퇴근 누락"], errors='coerce').fillna(0)

    print(f"   ✅ 기본 열 계산 완료\n")
    return df


def calculate_g_column(df: pd.DataFrame) -> pd.DataFrame:
    """G열 계산: 실제 근로시간(결근,퇴근누락포함).

    Args:
        df: 계산 중인 DataFrame

    Returns:
        G열이 추가된 DataFrame
    """
    print("🔢 G열 계산 중 (실제 근로시간, 결근/퇴근누락포함)...")

    df["G_실제근로시간_결근포함"] = (
        df["F_실제근로시간"] + df["결근"] * 8 + df["퇴근누락"] * 8
    ).round(2)

    print(f"   ✅ G열 계산 완료\n")
    return df


def calculate_h_column(
    df: pd.DataFrame, df1: pd.DataFrame, df2: pd.DataFrame
) -> pd.DataFrame:
    """H열 계산: 실제 근로시간(결근,퇴근누락포함) 실제퇴근시간-출근등록시간.

    SUMPRODUCT 로직:
    =SUMPRODUCT(
        (shiftee_data2!M$5:M$5000<>"") *
        ((INT(M*1440)/1440 - INT(I*1440)/1440) * 24 - S*24) *
        (shiftee_data2!B$5:B$5000=B10)
    ) + shiftee_data!W10*8 + shiftee_data!X10*8

    Args:
        df: 계산 중인 DataFrame
        df1: shiftee_data1 DataFrame
        df2: shiftee_data2 DataFrame

    Returns:
        H열이 추가된 DataFrame
    """
    print("🔢 H열 계산 중 (SUMPRODUCT 로직, 실제퇴근-출근)...")

    # df2에서 필요한 열 추출
    # B열: 이름, I열: 근무일정 시작시간, M열: 퇴근시간, S열: 총 휴게시간
    df2_filtered = df2[["이름", "근무일정\n시작시간", "퇴근시간", "(실제)\n총 휴게시간"]].copy()
    df2_filtered.columns = ["직원명", "시작시간", "퇴근시간", "휴게시간"]

    # M열(퇴근시간)이 비어있지 않은 행만 필터링
    df2_filtered = df2_filtered[df2_filtered["퇴근시간"].notna()]

    # 시간 계산 함수
    def calculate_work_hours(row):
        """각 행의 근무시간 계산 - Excel SUMPRODUCT 로직 구현"""
        from datetime import datetime, time, timedelta

        try:
            start = row["시작시간"]
            end = row["퇴근시간"]
            rest = row["휴게시간"]

            # 1. 출근/퇴근 시간을 Excel serial date 형식으로 변환
            # Excel serial date: 1.0 = 1 day, 0.5 = 12 hours (noon)
            def to_excel_serial(dt_value):
                """datetime/float를 Excel serial date로 변환"""
                if isinstance(dt_value, datetime):
                    # Excel epoch: 1899-12-30 (Excel's bug: 1900 is not a leap year)
                    excel_epoch = datetime(1899, 12, 30)
                    delta = dt_value - excel_epoch
                    return delta.total_seconds() / 86400  # 일 단위로 변환
                elif isinstance(dt_value, (int, float)):
                    return dt_value  # 이미 Excel serial date
                else:
                    return 0

            start_serial = to_excel_serial(start)
            end_serial = to_excel_serial(end)

            # 2. Excel의 INT(time*1440)/1440 로직 적용 (분 단위 절삭)
            # time*1440 = 분으로 변환, INT로 절삭, /1440으로 다시 일 단위
            start_minutes = int(start_serial * 1440)
            start_day = start_minutes / 1440

            end_minutes = int(end_serial * 1440)
            end_day = end_minutes / 1440

            # 3. 휴게시간 처리 (time 객체 또는 float)
            rest_hours = 0
            if rest is not None:
                if isinstance(rest, time):
                    # time 객체를 시간으로 변환
                    rest_hours = rest.hour + rest.minute / 60 + rest.second / 3600
                elif isinstance(rest, timedelta):
                    # timedelta를 시간으로 변환
                    rest_hours = rest.total_seconds() / 3600
                elif isinstance(rest, (int, float)):
                    # Excel serial date (일 단위) -> 시간 단위로 변환
                    rest_hours = rest * 24
                else:
                    rest_hours = 0

            # 4. Excel 수식 그대로 계산: (end - start) * 24 - rest * 24
            work_hours = (end_day - start_day) * 24 - rest_hours

            return max(0, work_hours)  # 음수 방지

        except Exception as e:
            # 디버깅용 에러 출력 (필요시)
            # print(f"Error calculating work hours: {e}, start={start}, end={end}, rest={rest}")
            return 0

    df2_filtered["근무시간"] = df2_filtered.apply(calculate_work_hours, axis=1)

    # 직원별 근무시간 합산
    employee_hours = df2_filtered.groupby("직원명")["근무시간"].sum().to_dict()

    # H열 계산: 직원별 합산 + 결근*8 + 퇴근누락*8
    def get_h_value(row):
        employee = row["B_직원"]
        sumproduct_hours = employee_hours.get(employee, 0)
        absence_hours = row["결근"] * 8 + row["퇴근누락"] * 8
        return sumproduct_hours + absence_hours

    df["H_실제근로시간_퇴근출근기반"] = df.apply(get_h_value, axis=1).round(2)

    print(f"   ✅ H열 계산 완료 (SUMPRODUCT)\n")
    return df


def calculate_i_j_l_columns(df: pd.DataFrame) -> pd.DataFrame:
    """I, J, L열 계산.

    I열: 표준 근로시간 = 표준근로시간 + 표준최대잔여유급시간
    J열: 표준 근로시간(결근,퇴근누락포함) = 표준근로시간 + 결근*8 + 퇴근누락*8 + 표준최대잔여유급시간
    L열: 법정근로시간 = D열 - K열 (소정근로시간 - 유급휴가시간)

    Args:
        df: 계산 중인 DataFrame

    Returns:
        I, J, L열이 추가된 DataFrame
    """
    print("🔢 I, J, L열 계산 중...")

    # I열: 표준 근로시간
    df["I_표준근로시간"] = (df["표준근로시간"] + df["표준최대잔여유급시간"]).round(2)

    # J열: 표준 근로시간(결근,퇴근누락포함)
    df["J_표준근로시간_결근포함"] = (
        df["표준근로시간"]
        + df["결근"] * 8
        + df["퇴근누락"] * 8
        + df["표준최대잔여유급시간"]
    ).round(2)

    # L열: 법정근로시간 = D - K (소정근로시간 - 유급휴가시간)
    df["L_법정근로시간"] = (df["D_소정근로시간"] - df["K_유급휴가시간"]).round(2)

    print(f"   ✅ I, J, L열 계산 완료\n")
    return df


def calculate_overtime_columns(df: pd.DataFrame) -> pd.DataFrame:
    """초과근로 관련 열 계산 (M, N, O, P).

    M열: 실제 초과근로시간 = IF((F-L)<0, 0, F-L)
    N열: 실제 초과근로시간(결근,퇴근누락포함) = IF((G-L)<0, 0, G-L)
    O열: 실제 초과근로시간(결근,퇴근누락포함,조기출근제외)
         = IF(OR((H-L)>N, H>300), N, IF((H-L)<0, 0, H-L))
    P열: 조기출근 합산 = N - O

    Args:
        df: 계산 중인 DataFrame

    Returns:
        M, N, O, P열이 추가된 DataFrame
    """
    print("🔢 초과근로 열 계산 중 (M, N, O, P)...")

    # M열: 실제 초과근로시간 = IF((F-L)<0, 0, F-L)
    df["M_실제초과근로시간"] = df.apply(
        lambda row: max(0, row["F_실제근로시간"] - row["L_법정근로시간"]), axis=1
    ).round(2)

    # N열: 실제 초과근로시간(결근,퇴근누락포함)
    df["N_실제초과근로_결근포함"] = df.apply(
        lambda row: max(0, row["G_실제근로시간_결근포함"] - row["L_법정근로시간"]),
        axis=1,
    ).round(2)

    # O열: 실제 초과근로시간(결근,퇴근누락포함,조기출근제외)
    def calculate_o(row):
        h_minus_l = row["H_실제근로시간_퇴근출근기반"] - row["L_법정근로시간"]
        n = row["N_실제초과근로_결근포함"]
        h = row["H_실제근로시간_퇴근출근기반"]

        if h_minus_l > n or h > 300:
            return n
        elif h_minus_l < 0:
            return 0
        else:
            return h_minus_l

    df["O_실제초과근로_조기출근제외"] = df.apply(calculate_o, axis=1).round(2)

    # P열: 조기출근 합산
    df["P_조기출근합산"] = (df["N_실제초과근로_결근포함"] - df["O_실제초과근로_조기출근제외"]).round(2)

    print(f"   ✅ 초과근로 열 계산 완료\n")
    return df


def calculate_legal_limits_and_risk(
    df: pd.DataFrame, start_date: datetime, end_date: datetime
) -> pd.DataFrame:
    """법규 기준 및 적정성 판정 (Q, R, S, T, U, V).

    Q열: 법정 초과 근로시간 = DAY(종료일)/7*10
    R열: 법규 위반(전일까지) = DAY(종료일)/7*12
    S열: 월법규 위반시간 = 12*4.3
    T열: 월말까지 가능한 초과근로시간 = IF((S-O)<0, "가능시간없음", S-O)
    U열: 적정성 = IF(O>Q, "위험", "정상")
    V열: 법규 기준초과자 = IF(AND(O<>0, O>=R), "법기준초과", "")

    Args:
        df: 계산 중인 DataFrame
        start_date: 시작 날짜
        end_date: 종료 날짜

    Returns:
        Q~V열이 추가된 DataFrame
    """
    print(f"🔢 법규 기준 및 적정성 판정 중 ({start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')})...")

    # Excel X1 셀: 분석 종료일의 날짜 (예: 10월 30일 → 30, 12월 15일 → 15)
    last_day = end_date.day
    print(f"   📅 기준일 (분석 종료): {last_day}일 ({end_date.year}년 {end_date.month}월 {end_date.day}일)")

    # Q열: 법정 초과 근로시간
    q_value = round(last_day / 7 * 10, 2)
    df["Q_법정초과근로시간"] = q_value
    print(f"   ⚖️  법정 초과근로시간: {q_value:.2f}시간")

    # R열: 법규 위반(전일까지)
    r_value = round(last_day / 7 * 12, 2)
    df["R_법규위반_전일까지"] = r_value
    print(f"   ⚠️  법규 위반(전일까지): {r_value:.2f}시간")

    # S열: 월법규 위반시간
    s_value = round(12 * 4.3, 2)
    df["S_월법규위반시간"] = s_value
    print(f"   🚨 월법규 위반시간: {s_value}시간")

    # T열: 월말까지 가능한 초과근로시간
    df["T_월말가능초과근로"] = df.apply(
        lambda row: (
            "가능시간없음"
            if (row["S_월법규위반시간"] - row["O_실제초과근로_조기출근제외"]) < 0
            else round(row["S_월법규위반시간"] - row["O_실제초과근로_조기출근제외"], 2)
        ),
        axis=1,
    )

    # U열: 적정성 (핵심!)
    df["U_적정성"] = df.apply(
        lambda row: (
            "위험" if row["O_실제초과근로_조기출근제외"] > row["Q_법정초과근로시간"] else "정상"
        ),
        axis=1,
    )

    # V열: 법규 기준초과자
    df["V_법규기준초과자"] = df.apply(
        lambda row: (
            "법기준초과"
            if (
                row["O_실제초과근로_조기출근제외"] != 0
                and row["O_실제초과근로_조기출근제외"] >= row["R_법규위반_전일까지"]
            )
            else ""
        ),
        axis=1,
    )

    print(f"   ✅ 법규 기준 및 적정성 판정 완료\n")
    return df


def print_summary(df: pd.DataFrame):
    """계산 결과 요약 출력.

    Args:
        df: 최종 계산 결과 DataFrame
    """
    print("\n" + "=" * 80)
    print("📊 적정성 분석 결과")
    print("=" * 80)

    total = len(df)
    risk_count = len(df[df["U_적정성"] == "위험"])
    normal_count = total - risk_count
    legal_exceed_count = len(df[df["V_법규기준초과자"] == "법기준초과"])

    print(f"\n총 직원: {total}명")
    print(f"  - ✅ 정상: {normal_count}명 ({normal_count/total*100:.1f}%)")
    print(f"  - ⚠️  위험: {risk_count}명 ({risk_count/total*100:.1f}%)")
    print(f"  - 🚨 법규기준초과: {legal_exceed_count}명 ({legal_exceed_count/total*100:.1f}%)")


def print_risk_employees(df: pd.DataFrame):
    """위험 직원 목록 출력.

    Args:
        df: 최종 계산 결과 DataFrame
    """
    risk_df = df[df["U_적정성"] == "위험"].copy()

    if risk_df.empty:
        print("\n✅ 위험 직원이 없습니다!\n")
        return

    print("\n" + "=" * 80)
    print("⚠️  위험 직원 목록")
    print("=" * 80)

    # 출력용 컬럼 선택
    display_columns = [
        "B_직원",
        "C_본조직",
        "O_실제초과근로_조기출근제외",
        "Q_법정초과근로시간",
        "U_적정성",
        "V_법규기준초과자",
    ]

    display_df = risk_df[display_columns].copy()
    display_df.columns = ["직원", "본조직", "실제초과근로(h)", "법정기준(h)", "적정성", "법규초과"]

    # 실제초과근로 시간으로 정렬 (내림차순)
    display_df = display_df.sort_values("실제초과근로(h)", ascending=False)

    print(display_df.to_string(index=False))
    print()


def save_to_excel(df: pd.DataFrame, output_path: Path, start_date: datetime, end_date: datetime):
    """결과를 Excel 파일로 저장하고 스타일링 적용.

    Args:
        df: 최종 계산 결과 DataFrame
        output_path: 저장할 Excel 파일 경로
        start_date: 시작 날짜
        end_date: 종료 날짜
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # 컬럼명을 원본 레포트 형식으로 변경
    column_mapping = {
        "B_직원": "직원",
        "C_본조직": "본조직",
        "D_소정근로시간": "소정\n근로시간",
        "E_승인된근로시간": "승인된\n근로시간",
        "F_실제근로시간": "실제\n근로시간",
        "G_실제근로시간_결근포함": "실제\n근로시간\n(결근,퇴근누락\n포함)",
        "H_실제근로시간_퇴근출근기반": "실제\n근로시간\n(퇴근출근기반)",
        "I_표준근로시간": "표준\n근로시간",
        "J_표준근로시간_결근포함": "표준\n근로시간\n(결근포함)",
        "K_유급휴가시간": "유급휴가\n시간",
        "L_법정근로시간": "법정\n근로시간",
        "M_실제초과근로시간": "실제 초과\n근로시간",
        "N_실제초과근로_결근포함": "실제 초과\n근로시간\n(결근포함)",
        "O_실제초과근로_조기출근제외": "실제 초과\n근로시간\n(조기출근제외)",
        "P_조기출근합산": "조기출근\n합산",
        "Q_법정초과근로시간": "법정\n초과 근로시간",
        "R_법규위반_전일까지": "법규 위반\n(전일까지)",
        "S_월법규위반시간": "월법규\n위반시간",
        "T_월말가능초과근로": "월말까지\n가능한\n초과근로시간",
        "U_적정성": "적정성",
        "V_법규기준초과자": "법규 기준초과자",
    }

    # 필요한 열만 선택하고 이름 변경
    display_columns = [
        "B_직원", "C_본조직",
        "D_소정근로시간", "E_승인된근로시간", "F_실제근로시간",
        "G_실제근로시간_결근포함", "H_실제근로시간_퇴근출근기반",
        "I_표준근로시간", "J_표준근로시간_결근포함",
        "K_유급휴가시간", "L_법정근로시간",
        "M_실제초과근로시간", "N_실제초과근로_결근포함",
        "O_실제초과근로_조기출근제외", "P_조기출근합산",
        "Q_법정초과근로시간", "R_법규위반_전일까지",
        "S_월법규위반시간", "T_월말가능초과근로",
        "U_적정성", "V_법규기준초과자"
    ]

    display_df = df[display_columns].copy()
    display_df = display_df.rename(columns=column_mapping)

    # A열 (빈 열) 추가
    display_df.insert(0, "", "")

    # DataFrame을 Excel로 저장
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        display_df.to_excel(writer, sheet_name="계산", index=False)

        # 워크시트 가져오기
        ws = writer.sheets["계산"]

        # W, X열 추가 (원본 레포트와 동일하게)
        w_col = len(display_df.columns) + 1  # 빈 열 포함한 전체 컬럼 수 + 1
        x_col = w_col + 1

        ws.cell(row=1, column=w_col).value = 167
        ws.cell(row=1, column=x_col).value = end_date

        for row_idx in range(2, len(display_df) + 2):
            ws.cell(row=row_idx, column=w_col).value = 167
            ws.cell(row=row_idx, column=x_col).value = end_date

        # 스타일 정의
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_font = Font(bold=True, size=11)
        risk_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        legal_exceed_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 헤더 스타일 적용 (1행 전체, A열 포함)
        for col_idx in range(1, x_col + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        # 적정성, 법규초과 컬럼 인덱스 찾기 (display_df 기준)
        adequacy_col_idx = display_df.columns.get_loc("적정성") + 1  # Excel은 1-based
        legal_exceed_col_idx = display_df.columns.get_loc("법규 기준초과자") + 1

        # 데이터 행 스타일 적용 (2행부터)
        for row_idx in range(2, len(display_df) + 2):
            adequacy = ws.cell(row=row_idx, column=adequacy_col_idx).value
            legal_exceed = ws.cell(row=row_idx, column=legal_exceed_col_idx).value

            # 위험 직원은 빨간색 배경
            if adequacy == "위험":
                for col_idx in range(1, x_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = risk_fill
                    cell.border = border
            # 법규 기준 초과자는 주황색 배경
            elif legal_exceed == "법기준초과":
                for col_idx in range(1, x_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = legal_exceed_fill
                    cell.border = border
            else:
                # 일반 직원은 테두리만 적용
                for col_idx in range(1, x_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = border

            # 숫자 열은 가운데 정렬
            for col_idx in range(1, x_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.alignment = center_align

        # 열 너비 자동 조정
        ws.column_dimensions["A"].width = 5  # A열은 좁게
        ws.column_dimensions[get_column_letter(w_col)].width = 8  # W열
        ws.column_dimensions[get_column_letter(x_col)].width = 12  # X열

        # 나머지 컬럼 너비 자동 조정
        for col_idx in range(2, w_col):  # B열부터 V열까지
            col_letter = get_column_letter(col_idx)
            header_value = ws.cell(row=1, column=col_idx).value
            max_length = len(str(header_value)) if header_value else 10

            # 데이터 길이 확인 (최대 10행만 샘플링)
            for row_idx in range(2, min(12, len(display_df) + 2)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            # 최소 10, 최대 30으로 제한
            adjusted_width = min(max(max_length + 2, 10), 30)
            ws.column_dimensions[col_letter].width = adjusted_width

        # 틀 고정 (헤더 행 고정, B열부터)
        ws.freeze_panes = "B2"

    print(f"📊 Excel 리포트 생성 완료: {output_path}")
    print(f"   기간: {start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}")
    print(f"   파일 크기: {output_path.stat().st_size / 1024:.1f} KB")

    # 요약 정보
    total = len(df)
    risk_count = len(df[df["U_적정성"] == "위험"])
    legal_exceed_count = len(df[df["V_법규기준초과자"] == "법기준초과"])

    print(f"   직원 수: {total}명 (위험: {risk_count}명, 법규초과: {legal_exceed_count}명)\n")


def main():
    """메인 함수."""
    import argparse

    parser = argparse.ArgumentParser(
        description="Template 없이 직접 계산하는 적정성 위험 판정",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
예제:
  # 기본 실행 (data/ 디렉토리의 shiftee_data1.xlsx, shiftee_data2.xlsx 사용)
  python scripts/calculate_risk_direct.py

  # 특정 파일 지정
  python scripts/calculate_risk_direct.py --data1 data/shiftee_data1.xlsx --data2 data/shiftee_data2.xlsx

  # CSV로 저장
  python scripts/calculate_risk_direct.py --output risk_report.csv

  # Excel 리포트로 저장 (스타일링 포함)
  python scripts/calculate_risk_direct.py --output risk_report.xlsx

  # 특정 기간 지정 (법규 기준 계산용)
  python scripts/calculate_risk_direct.py --start 2025-11-01 --end 2025-11-30 --output report_11월.xlsx
  python scripts/calculate_risk_direct.py --start 2025-12-01 --end 2025-12-14 --output report_12월.xlsx

  # 다운로드부터 분석까지 한 번에 실행하고 Excel 리포트 생성
  python scripts/calculate_risk_direct.py --download --start 2025-11-01 --end 2025-11-30 --output report.xlsx

  # 분석 후 위험 직원 목록을 카카오톡으로 전송
  python scripts/calculate_risk_direct.py --send-kakao

  # 요약 메시지만 카카오톡으로 전송
  python scripts/calculate_risk_direct.py --kakao-summary

  # 다운로드 + 분석 + 카카오톡 전송을 한 번에
  python scripts/calculate_risk_direct.py --download --send-kakao
        """,
    )
    parser.add_argument(
        "--data1",
        type=Path,
        default=Path("data/shiftee_data1.xlsx"),
        help="shiftee_data1.xlsx 파일 경로 (기본: data/shiftee_data1.xlsx)",
    )
    parser.add_argument(
        "--data2",
        type=Path,
        default=Path("data/shiftee_data2.xlsx"),
        help="shiftee_data2.xlsx 파일 경로 (기본: data/shiftee_data2.xlsx)",
    )
    parser.add_argument(
        "--output",
        "-o",
        type=Path,
        help="결과를 CSV 또는 Excel 파일로 저장 (기본: output/report_YYYYMMDD.xlsx)",
    )
    parser.add_argument(
        "--start",
        type=str,
        help="시작 날짜 (YYYY-MM-DD 형식, 기본: 현재 월 1일)",
    )
    parser.add_argument(
        "--end",
        type=str,
        help="종료 날짜 (YYYY-MM-DD 형식, 기본: 전일(어제))",
    )
    parser.add_argument(
        "--download",
        action="store_true",
        help="Shiftee에서 데이터를 먼저 다운로드한 후 분석 (날짜 미지정시 기본값 사용)",
    )
    parser.add_argument(
        "--send-kakao",
        action="store_true",
        help="분석 완료 후 위험 직원 목록을 카카오톡으로 전송",
    )
    parser.add_argument(
        "--kakao-summary",
        action="store_true",
        help="카카오톡으로 요약 메시지만 전송 (위험 직원 목록 제외)",
    )

    args = parser.parse_args()

    # 출력 파일 기본값 설정 (output/report_YYYYMMDD.xlsx)
    if not args.output:
        today = datetime.now().strftime("%Y%m%d")
        args.output = Path("output") / f"report_{today}.xlsx"

    # 날짜 파싱 및 기본값 설정
    now = datetime.now()

    # --start 처리: 미지정시 현재 월의 1일
    if args.start:
        try:
            start_date = datetime.strptime(args.start, "%Y-%m-%d")
        except ValueError as e:
            print(f"❌ 시작 날짜 형식 오류: {e}")
            print("💡 올바른 형식: YYYY-MM-DD (예: 2025-11-01)")
            return 1
    else:
        # 기본값: 현재 월의 1일
        start_date = datetime(now.year, now.month, 1)
        print(f"📅 --start 미지정 → 이번 달 1일 사용: {start_date.strftime('%Y-%m-%d')}")

    # --end 처리: 미지정시 가장 최근 평일 (주말 제외)
    if args.end:
        try:
            end_date = datetime.strptime(args.end, "%Y-%m-%d")
        except ValueError as e:
            print(f"❌ 종료 날짜 형식 오류: {e}")
            print("💡 올바른 형식: YYYY-MM-DD (예: 2025-11-30)")
            return 1
    else:
        # 기본값: 무조건 전일 (어제)
        end_date = now - timedelta(days=1)
        end_date = end_date.replace(hour=0, minute=0, second=0, microsecond=0)
        print(f"📅 --end 미지정 → 전일 사용: {end_date.strftime('%Y-%m-%d')}")

    if not args.start and not args.end:
        print()

    try:
        # 다운로드 실행 (옵션이 있을 경우)
        if args.download:
            output_dir = args.data1.parent  # data/ 디렉토리
            output_dir.mkdir(parents=True, exist_ok=True)

            # 비동기 다운로드 실행
            data1_path, data2_path = asyncio.run(
                download_shiftee_data(start_date, end_date, output_dir)
            )

            print("=" * 80)
            print("✅ 다운로드 완료")
            print("=" * 80)
            print(f"  - {data1_path}")
            print(f"  - {data2_path}")
            print()

            # 다운로드된 파일로 경로 업데이트
            args.data1 = data1_path
            args.data2 = data2_path
        print("=" * 80)
        print("🚀 Template 없이 직접 계산하는 적정성 위험 판정")
        print("=" * 80)
        print()

        # 1. 데이터 로드
        print("📂 1단계: 데이터 로드")
        print("-" * 80)
        df1 = load_shiftee_data1(args.data1)
        df2 = load_shiftee_data2(args.data2)

        # 1-1. 교대제 직원 제외
        if "본직무" in df1.columns:
            original_count = len(df1)
            shift_workers = df1[df1["본직무"] == "교대제"]
            shift_worker_names = shift_workers["직원"].tolist() if "직원" in shift_workers.columns else []

            df1 = df1[df1["본직무"] != "교대제"].copy()
            excluded_count = original_count - len(df1)

            if excluded_count > 0:
                print(f"\n⚠️  교대제 직원 {excluded_count}명 제외:")
                for name in shift_worker_names:
                    print(f"   - {name}")
                print()

                # df2에서도 해당 직원들의 출퇴근 기록 제외
                if "이름" in df2.columns:
                    df2 = df2[~df2["이름"].isin(shift_worker_names)].copy()
                    print(f"   ✅ 출퇴근 기록에서도 제외 완료\n")

        # 2. 단계별 계산
        print("🔢 2단계: 계산 수행")
        print("-" * 80)

        df = calculate_basic_columns(df1)
        df = calculate_g_column(df)
        df = calculate_h_column(df, df1, df2)
        df = calculate_i_j_l_columns(df)
        df = calculate_overtime_columns(df)
        df = calculate_legal_limits_and_risk(df, start_date, end_date)

        # 3. 결과 출력
        print_summary(df)
        print_risk_employees(df)

        # 4. 파일 저장 (CSV 또는 Excel)
        # 출력 디렉터리 생성
        args.output.parent.mkdir(parents=True, exist_ok=True)

        if args.output.suffix.lower() == ".xlsx":
            # Excel 파일로 저장
            save_to_excel(df, args.output, start_date, end_date)
        else:
            # CSV 파일로 저장
            df.to_csv(args.output, index=False, encoding="utf-8-sig")
            print(f"💾 CSV 저장 완료: {args.output}")
            print(f"   파일 크기: {args.output.stat().st_size / 1024:.1f} KB\n")

        # 5. 카카오톡 메시지 전송 (옵션)
        if args.send_kakao or args.kakao_summary:
            try:
                # kakao_send 모듈을 여기서 import (kakao_token.json이 없을 때 오류 방지)
                import os
                kakao_send_path = project_root / "kakao_send"
                if str(kakao_send_path) not in sys.path:
                    sys.path.insert(0, str(kakao_send_path))

                from kakao_send import Kakao, format_risk_message
                from kakao_send.message_formatter import format_summary_message

                print("=" * 80)
                print("📱 카카오톡 메시지 전송")
                print("=" * 80)

                # Kakao 인스턴스 생성
                kakao = Kakao()

                # 메시지 포맷 선택
                if args.kakao_summary:
                    message = format_summary_message(df, start_date, end_date)
                else:
                    message = format_risk_message(df, start_date, end_date, show_all=False)

                # 메시지 전송 (토큰 자동 갱신 포함)
                if kakao.send_message(message):
                    print()
                else:
                    print("⚠️  메시지 전송에 실패했습니다.\n")

            except FileNotFoundError as e:
                print(f"\n⚠️  카카오톡 설정이 필요합니다: {e}")
                print("\n📝 설정 방법:")
                print("   1. kakao_send/kakao_get_token.py 파일을 수정하여 초기 토큰 발급")
                print("   2. KAKAO_APP_KEY 환경변수 설정 또는 .env 파일에 추가")
                print("   3. 자세한 내용은 docs/KAKAO_SETUP.md 참조\n")
            except ImportError as e:
                print(f"\n⚠️  kakao_send 모듈을 불러올 수 없습니다: {e}")
                print("   requests 패키지가 설치되어 있는지 확인하세요: pip install requests\n")
            except Exception as e:
                print(f"\n⚠️  카카오톡 메시지 전송 중 오류 발생: {e}\n")

        return 0

    except FileNotFoundError as e:
        print(f"\n❌ 파일을 찾을 수 없습니다: {e}\n")
        return 1
    except Exception as e:
        print(f"\n❌ 오류 발생: {e}", file=sys.stderr)
        import traceback

        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
