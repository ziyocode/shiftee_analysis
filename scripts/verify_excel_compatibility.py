#!/usr/bin/env python3
"""
Excel 호환성 검증 스크립트

생성된 Excel 파일이 모든 호환성 요구사항을 만족하는지 확인합니다.
"""

import sys
import zipfile
from pathlib import Path


def verify_excel_file(file_path: Path) -> bool:
    """Excel 파일 호환성 검증.

    Args:
        file_path: 검증할 Excel 파일 경로

    Returns:
        검증 통과 여부
    """
    print(f"🔍 Excel 파일 호환성 검증: {file_path.name}\n")

    if not file_path.exists():
        print(f"❌ 파일을 찾을 수 없습니다: {file_path}")
        return False

    # 파일 크기 확인
    file_size_kb = file_path.stat().st_size / 1024
    print(f"📏 파일 크기: {file_size_kb:.1f} KB")

    try:
        # ZIP 구조 검증
        print(f"\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        print(f"📦 ZIP 구조 검증")
        print(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")

        with zipfile.ZipFile(file_path, "r") as zf:
            files = zf.namelist()
            print(f"✅ ZIP 파일 개수: {len(files)}")

            # 문제가 될 수 있는 파일 확인
            external_links = [f for f in files if "externalLink" in f]
            vba_files = [f for f in files if "vba" in f.lower()]
            calc_chain = [f for f in files if "calcChain" in f]

            issues = []

            # 외부 링크 확인
            print(f"\n🔗 외부 링크 확인:")
            if external_links:
                print(f"   ⚠️  발견됨: {len(external_links)}개")
                for link in external_links:
                    print(f"      - {link}")
                issues.append("외부 링크가 존재합니다")
            else:
                print(f"   ✅ 없음 (Excel 호환성 양호)")

            # VBA 파일 확인
            print(f"\n📜 VBA 매크로 확인:")
            if vba_files:
                print(f"   ⚠️  발견됨: {len(vba_files)}개")
                for vba in vba_files:
                    print(f"      - {vba}")
                issues.append("VBA 파일이 존재합니다 (macOS 호환성 주의)")
            else:
                print(f"   ✅ 없음 (macOS Excel 호환성 양호)")

            # CalcChain 확인
            print(f"\n🧮 CalcChain 확인:")
            if calc_chain:
                print(f"   ⚠️  발견됨: {len(calc_chain)}개")
                for chain in calc_chain:
                    print(f"      - {chain}")
                issues.append("CalcChain이 존재합니다 (호환성 주의)")
            else:
                print(f"   ✅ 없음 (Excel 호환성 양호)")

        # openpyxl 검증
        print(f"\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        print(f"🔬 openpyxl 검증")
        print(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")

        import openpyxl

        wb = openpyxl.load_workbook(file_path)
        print(f"✅ 파일을 열 수 있습니다")

        print(f"\n📋 시트 구조:")
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(
                f"   - '{sheet_name}': {sheet.max_row:,} rows × {sheet.max_column} cols"
            )

        # 계산 모드 확인
        if hasattr(wb, "calculation"):
            print(f"\n🧮 계산 설정:")
            print(f"   - 계산 모드: {wb.calculation.calcMode}")
            print(f"   - 로드 시 전체 계산: {wb.calculation.fullCalcOnLoad}")

        wb.close()

        # 최종 결과
        print(f"\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        print(f"📊 검증 결과")
        print(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")

        if issues:
            print(f"\n⚠️  발견된 문제점:")
            for i, issue in enumerate(issues, 1):
                print(f"   {i}. {issue}")
            print(f"\n💡 권장 조치:")
            print(f"   - 파일을 다시 생성하여 fresh workbook 방식으로 저장하세요")
            print(f"   - use_fresh_workbook=True로 설정 확인")
            return False
        else:
            print(f"\n✅ 모든 호환성 검사를 통과했습니다!")
            print(f"\n📱 다음 단계:")
            print(f"   1. Excel에서 파일을 열어보세요")
            print(f"   2. 수식이 자동으로 계산되는지 확인하세요")
            print(f"   3. 모든 시트가 정상적으로 표시되는지 확인하세요")
            return True

    except Exception as e:
        print(f"\n❌ 검증 중 오류 발생: {e}")
        import traceback

        traceback.print_exc()
        return False


def main():
    """메인 함수."""
    import argparse

    parser = argparse.ArgumentParser(
        description="Excel 파일 호환성 검증",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
예제:
  python verify_excel_compatibility.py output/report.xlsx
  python verify_excel_compatibility.py output/report_final_fixed.xlsx
        """,
    )
    parser.add_argument("file", type=Path, help="검증할 Excel 파일 경로")

    args = parser.parse_args()

    success = verify_excel_file(args.file)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
