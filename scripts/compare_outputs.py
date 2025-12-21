#!/usr/bin/env python
"""Compare generated report with reference."""

import openpyxl

# Load generated report
wb = openpyxl.load_workbook("test_report.xlsx", data_only=True)
ws = wb.active

for row in range(2, ws.max_row + 1):
    name = ws.cell(row=row, column=2).value
    if name == "이상훈":
        print(f"Generated 이상훈 (Row {row}):")
        print(f"  B (직원): {ws.cell(row=row, column=2).value}")
        print(f"  F (실제): {ws.cell(row=row, column=6).value:.2f}")
        print(f"  H (SUMPRODUCT): {ws.cell(row=row, column=13).value:.2f}")
        print(f"  L (법정): {ws.cell(row=row, column=16).value:.2f}")
        print(f"  O (조기출근제외): {ws.cell(row=row, column=19).value:.2f}")
        print(f"  Q (법정기준): {ws.cell(row=row, column=21).value:.2f}")
        print(f"  U (적정성): {ws.cell(row=row, column=25).value}")
        print(f"  W (167): {ws.cell(row=row, column=27).value}")
        print(f"  X (date): {ws.cell(row=row, column=28).value}")

        # Check fill color
        cell = ws.cell(row=row, column=2)
        if cell.fill and cell.fill.start_color:
            rgb = cell.fill.start_color.rgb
            is_red = rgb == "00FFE6E6"
            print(f"  Fill color: {rgb} ({'✅ red' if is_red else '❌ other'})")
        break

print("\n=== Reference File ===")
wb_ref = openpyxl.load_workbook("data/레포트_20251101-1130_뱅킹인프라본부.xlsx", data_only=True)
ws_ref = wb_ref.active

for row in range(2, ws_ref.max_row + 1):
    name = ws_ref.cell(row=row, column=2).value
    if name == "이상훈":
        print(f"Reference 이상훈 (Row {row}):")
        print(f"  F (실제): {ws_ref.cell(row=row, column=6).value:.2f}")
        print(f"  H (SUMPRODUCT): {ws_ref.cell(row=row, column=8).value:.2f}")
        print(f"  L (법정): {ws_ref.cell(row=row, column=12).value:.2f}")
        print(f"  O (조기출근제외): {ws_ref.cell(row=row, column=15).value:.2f}")
        print(f"  Q (법정기준): {ws_ref.cell(row=row, column=17).value:.2f}")
        print(f"  U (적정성): {ws_ref.cell(row=row, column=21).value}")
        break

print("\n=== Differences ===")
# Calculate differences
gen_vals = {
    'F': ws.cell(row=14, column=6).value,
    'H': ws.cell(row=14, column=13).value,
    'L': ws.cell(row=14, column=16).value,
    'O': ws.cell(row=14, column=19).value,
    'Q': ws.cell(row=14, column=21).value,
}

# Find reference row
ref_row = None
for row in range(2, ws_ref.max_row + 1):
    if ws_ref.cell(row=row, column=2).value == "이상훈":
        ref_row = row
        break

if ref_row:
    ref_vals = {
        'F': ws_ref.cell(row=ref_row, column=6).value,
        'H': ws_ref.cell(row=ref_row, column=8).value,
        'L': ws_ref.cell(row=ref_row, column=12).value,
        'O': ws_ref.cell(row=ref_row, column=15).value,
        'Q': ws_ref.cell(row=ref_row, column=17).value,
    }

    for key in gen_vals:
        diff = abs(gen_vals[key] - ref_vals[key])
        status = "✅" if diff < 0.01 else "❌"
        print(f"{key}: {status} diff={diff:.6f}")
